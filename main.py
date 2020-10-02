import os
import re
import subprocess
import sys
from datetime import datetime
from functools import partial
from time import sleep

from PIL import Image, ImageDraw, ImageFont, ImageGrab
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import QThread
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QCompleter, QTableWidgetItem, QDesktopWidget, QMainWindow, QWidget
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver

# from interface import *
from interface import Ui_MainWindow


class SearchThread(QThread):
    running = False

    def __init__(self, mainwindow):
        QThread.__init__(self)
        self.mainwindow = mainwindow
        self.running = True

    def run(self):
        # Запуск функции поиска
        while self.running:
            self.mainwindow.start_searching()
        self.running = True

    def stop(self):
        self.mainwindow.ui.pushButton_Start.setDisabled(False)
        self.running = False


class Window(QWidget):
    def __init__(self):
        super(Window, self).__init__()
        self.setWindowIcon(QIcon('pic/error.png'))
        self.setWindowTitle('Внимание!')
        self.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)
        self.setGeometry(QtCore.QRect(500, 500, 650, 150))
        self.label_Warning = QtWidgets.QLabel(self)
        self.label_Warning.move(30, 30)
        self.label_Warning.setText("""Размер приложений и текста на вашем дисплее увеличены.
При работе программы, обводящая рамка объявлений на скриншотах, будет отображаться некорректно.
Во избежание данной ошибки установите значение в настройках Windows->Дисплей->Масштабирование 100%""")


class MyWin(QMainWindow):
    def __init__(self, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        # Иконка приложения
        self.setWindowIcon(QIcon('pic/logo.png'))

        # Логотип в правом нижнем углу
        self.lbl = QtWidgets.QLabel(self)
        self.pix = QtGui.QPixmap('pic/logo.png')
        self.lbl.setPixmap(self.pix)
        self.lbl.resize(500, 500)
        self.lbl.move(1280, 600)

        # Экземпляр потока
        self.thread_instance = SearchThread(self)

        # Настройки таблицы для Гугла
        self.ui.tableWidget_Google.setColumnCount(1)
        self.ui.tableWidget_Google.setRowCount(5)

        # Настройки таблицы для Яндекса
        self.ui.tableWidget.setColumnCount(2)
        self.ui.tableWidget.setRowCount(5)

        # Массив с регионами Гугла из эксель файла
        self.gl_regions = self.get_gl_regions()
        print(self.gl_regions)

        # Настройки завершателя слов для Гугла
        completer_gl = QCompleter([*self.gl_regions.keys()])
        completer_gl.setCaseSensitivity(False)

        # Массив с регионами Яндекса из эксель файла
        yd_regions = self.get_yd_regions()
        print(yd_regions)

        # Настройки завершателя слов для Яндекса
        completer_yd = QCompleter(yd_regions)
        completer_yd.setCaseSensitivity(False)

        # В ячейки таблицы первого столбца устанавливаем QLineEdit, в которые пользователь будет писать регионы
        line_edits = []
        for i in range(0, self.ui.tableWidget.rowCount()):
            line_edits.append(QtWidgets.QLineEdit())
            # Устанавливаем QLineEdit QCompleter, чтобы он завершал слова, которые пишет пользователь
            line_edits[i].setCompleter(completer_yd)
            self.ui.tableWidget.setCellWidget(i, 0, line_edits[i])
            # Передаём частями в функцию данные о тексте, о номере строки и столбца, в которой находится QLineEdit
            line_edits[i].editingFinished.connect(partial(self.get_lr, cell=self.ui.tableWidget.cellWidget(i, 0), row=i, column=1))

        print(line_edits)
        print(self.ui.tableWidget.rowCount())

        line_edits_gl = []
        for i in range(0, self.ui.tableWidget_Google.rowCount()):
            line_edits_gl.append(QtWidgets.QLineEdit())
            line_edits_gl[i].setCompleter(completer_gl)
            self.ui.tableWidget_Google.setCellWidget(i, 0, line_edits_gl[i])

        self.ui.textEdit_Requests.setToolTip("""Поле для ввода запросов.
Одна строка - один запрос.
Желательно без пустых строк.""")
        self.ui.textEdit_SitesAddresses.setToolTip("""Поле для ввода адресов сайтов.
Одна строка - один адрес.
Например: test.ru
Без http/https
Без www.
Без знаков после домена верхнего уровня.""")
        self.ui.radioButton_Windowscreen.setToolTip("""Скриншоты размером с экран с панелью пуск""")
        self.ui.radioButton_Fullscreen.setToolTip("""Скриншоты без панели пуск, но с полной страницей выдачи""")
        self.ui.radioButton_OnlyAd.setToolTip("""Скриншоты только рекламы спец искомого сайта и ничего лишнего""")
        self.ui.radioButton_SpecialAndGarant.setToolTip("""Скриншоты спец и гарант без определенного сайта""")
        self.ui.checkBox_AddTimeDateToScreen.setToolTip("""Имеет смысл только для fullscreen режима""")
        self.ui.tableWidget.setToolTip("""Таблица для заполнения регионами.
Если нет нужного города, то его можно добавить вручную
Зайти в яндекс, поменять в левом верхнем углу местоположение
Сделать любой тестовый запрос
В адресной стркое найти параметр lr= (это будет всегда число)
Вставить его в таблицу""")
        self.ui.tableWidget_Google.setToolTip("""Вводить на латинице""")
        self.ui.groupBox_SavePath.setToolTip("""Выберите область для сохранения результатов
Программа автоматически создаст папку, в которую поместит скриншоты""")

        # Переменные для
        self.searchers = []  # для хранения всех систем поиска
        self.save_path = None  # путь для сохранения

        # Назначаем кнопке Старт функцию start_search, которая отвечает за запуск метода run() в классе потока
        self.ui.pushButton_Start.clicked.connect(self.start_search)
        self.ui.pushButton_Cancel.clicked.connect(self.end_search)
        self.ui.pushButton_SavePath.clicked.connect(self.get_save_path)

        # Назначаем на клик по чекбоксу функции, которые добавляют/убирают из списка поиск в системах
        self.ui.checkBox_Google.clicked.connect(self.settings_google)
        self.ui.checkBox_Yandex.clicked.connect(self.settings_yandex)

        self.ui.radioButton_Windowscreen.toggled.connect(self.disable_tools)
        self.ui.radioButton_Fullscreen.toggled.connect(self.disable_tools)
        self.ui.radioButton_OnlyAd.toggled.connect(self.disable_tools)
        self.ui.radioButton_SpecialAndGarant.toggled.connect(self.disable_tools)

        # Значения элементов интерфейса по умолчанию
        self.ui.label_Info.setText('Нажмите "Начать" для выполнения программы')
        self.ui.tableWidget_Google.setDisabled(True)
        self.ui.tableWidget.setDisabled(True)
        self.ui.radioButton_Windowscreen.setChecked(True)

        # Заглушки, пока функционал не готов
        self.ui.pushButton_Cancel.setDisabled(True)
        self.ui.pushButton_Cancel.setVisible(False)
        self.ui.label_Warning.setVisible(False)
        # self.ui.checkBox_Numeration.setDisabled(True)
        # self.ui.checkBox_Numeration.setVisible(False)

         # Временный текст
#         self.ui.textEdit_Requests.setText("""ноутбук""")
#         self.ui.textEdit_SitesAddresses.setText("""citilink.ru
# mvideo.ru
# asus.com""")

        # Предупреждающее окно
        q = QDesktopWidget().availableGeometry()
        print(q.height())
        if q.height() < 1040:
            print(q.height())
            self.w = Window()
            self.w.show()

    # Ф-ия присваивает переменным путь, который задаёт пользователь
    def get_save_path(self):
        self.save_path = QtWidgets.QFileDialog.getExistingDirectory()
        self.ui.label_SavePath.setText('Путь: {0}'.format(self.save_path))

    # Функция добавляет/удаляет в/из списка url для поиска в яндексе
    def settings_yandex(self):
        if self.ui.checkBox_Yandex.isChecked():
            self.searchers.append('https://yandex.ru/search/?text={0}&lr={1}')
            self.ui.tableWidget.setDisabled(False)
        else:
            self.searchers.remove('https://yandex.ru/search/?text={0}&lr={1}')
            self.ui.tableWidget.setDisabled(True)
        print(self.searchers)

    # Функция добавляет/удаляет в/из списка url для поиска в гугле
    def settings_google(self):
        if self.ui.checkBox_Google.isChecked():
            self.searchers.append('https://www.google.com/search?q={0}&uule={1}')
            self.ui.tableWidget_Google.setDisabled(False)
        else:
            self.searchers.remove('https://www.google.com/search?q={0}&uule={1}')
            self.ui.tableWidget_Google.setDisabled(True)
        print(self.searchers)

    def disable_tools(self):
        if self.ui.radioButton_Windowscreen.isChecked():
            self.ui.checkBox_AddTimeDateToScreen.setDisabled(True)
            self.ui.checkBox_WithoutFrame.setDisabled(False)
            self.ui.checkBox_WithoutScrollDown.setDisabled(False)
            self.ui.checkBox_Numeration.setDisabled(False)
            self.ui.textEdit_SitesAddresses.setDisabled(False)
        elif self.ui.radioButton_Fullscreen.isChecked():
            self.ui.checkBox_AddTimeDateToScreen.setDisabled(False)
            self.ui.checkBox_WithoutFrame.setDisabled(False)
            self.ui.checkBox_WithoutScrollDown.setDisabled(True)
            self.ui.checkBox_Numeration.setDisabled(False)
            self.ui.textEdit_SitesAddresses.setDisabled(False)
        elif self.ui.radioButton_OnlyAd.isChecked():
            self.ui.checkBox_AddTimeDateToScreen.setDisabled(True)
            self.ui.checkBox_WithoutFrame.setDisabled(True)
            self.ui.checkBox_WithoutScrollDown.setDisabled(True)
            self.ui.checkBox_Numeration.setDisabled(True)
            self.ui.textEdit_SitesAddresses.setDisabled(False)
        elif self.ui.radioButton_SpecialAndGarant.isChecked():
            self.ui.checkBox_AddTimeDateToScreen.setDisabled(True)
            self.ui.checkBox_WithoutFrame.setDisabled(True)
            self.ui.checkBox_WithoutScrollDown.setDisabled(True)
            self.ui.checkBox_Numeration.setDisabled(True)
            self.ui.textEdit_SitesAddresses.setDisabled(True)


    # Функция разбивает текст пользователя в поле Запросы на элементы массива и возвращает массив
    def get_requests(self,text):
        temp = text.split('\n')
        user_requests = tuple(request for request in temp if request != '')
        return user_requests

    def get_sites_addresses(self, text):
        clear_sites_addresses = []
        pattern = r'\w+-*\w+\.\w+-*\w+\.*\w*'
        sites_addresses = text.split('\n')
        print(sites_addresses)

        for site in sites_addresses:
            match = re.search(pattern, site)
            if match:
                match = match[0]
                if match[0:4] == 'www.':
                    match = match[4:]
                clear_sites_addresses.append(match)
        print(clear_sites_addresses)
        clear_sites_addresses = tuple(site for site in clear_sites_addresses)
        return clear_sites_addresses

    # Функция вычисляет позицию сайта относительно блоков спец/сео/гарант
    def get_site_position(self, mas, site_address):
        count = 0
        for index, m in enumerate(mas, start=1):
            if site_address in m[1]:
                count = m[0]
                break
        if count == 0:
            index = 0

        return count, index

    def get_block_of_ads(self, results):
        special = []
        seo = []
        garant = []

        temp = False

        # Распределяем результаты по массивам
        for result in results:
            if 'Яндекс.Маркет' in result[1] and 'Реклама' in result[1]:
                continue
            if 'реклама' in result[1] or 'Реклама' in result[1]:
                if temp:
                    garant.append(result)
                else:
                    special.append(result)
            else:
                temp = True
                seo.append(result)

        block_of_ads = [special, seo, garant]
        return block_of_ads

    # Функция для определения позиций в спец/seo/гарант
    def get_positions(self, results, site_address=None):
        special = []
        seo = []
        garant = []

        temp = False

        # Распределяем результаты по массивам
        for result in results:
            if 'реклама' in result[1] or 'Реклама' in result[1]:
                if temp:
                    garant.append(result)
                else:
                    special.append(result)
            else:
                temp = True
                seo.append(result)

        special_position = self.get_site_position(special, site_address)
        seo_position = self.get_site_position(seo, site_address)
        garant_position = self.get_site_position(garant, site_address)

        positions = [special_position, seo_position, garant_position]
        block_of_ads = [special, seo, garant]
        return positions, block_of_ads

    # Функция возвращает скрин с нумерацией и рамками
    def edit_screen(self, screen_name, results, positions, block_of_ads):
        # Начинаем работу с изображением
        image = Image.open(screen_name)
        draw = ImageDraw.Draw(image)
        print('ща начну рисовать')
        # Настройки для ширфта
        font = ImageFont.truetype('Aegean.ttf', 25)

        # Рисуем дату и время
        if self.ui.checkBox_AddTimeDateToScreen.isChecked():
            date = datetime.now().strftime('%d.%m.%Y')
            time = datetime.now().strftime('%H:%M')
            draw.text((15, 60), time, fill=(255, 0, 0), font=font)
            draw.text((15, 95), date, fill=(255, 0, 0), font=font)

        print('нарисовал дату время')

        # Выделяем рамкой искомые запросы
        if not self.ui.checkBox_WithoutFrame.isChecked():
            for position in positions:
                if position[0] != 0:
                    element = results[position[0] - 1]
                    print(element[2])
                    print(element[3])
                    draw.rectangle((element[2]['x'], element[2]['y'], element[2]['x'] + element[3]['width'],
                                    element[2]['y'] + element[3]['height']), outline=(255, 0, 0, 255), width=3,)

        if self.ui.checkBox_Numeration.isChecked():
            # Рисуем на скрине нумерацию
            for result in results:
                draw.text((result[2]['x'] - 100, result[2]['y']), str(result[0]), fill=(0, 0, 0), font=font)

            for index, spcl in enumerate(block_of_ads[0], start=1):
                draw.text((spcl[2]['x'] - 50, spcl[2]['y']), str(index), fill=(255, 0, 0), font=font)

            for index, s in enumerate(block_of_ads[1], start=1):
                draw.text((s[2]['x'] - 50, s[2]['y']), str(index), fill=(0, 128, 0), font=font)

            for index, g in enumerate(block_of_ads[2], start=1):
                draw.text((g[2]['x'] - 50, g[2]['y']), str(index), fill=(0, 0, 255), font=font)

        # Конец рисования
        del draw
        image.save(screen_name)
        print('сохранил новый рисунок')

    # Функция возвращает значение lr для региона, который выбрал пользователь
    def get_lr(self, cell, row, column):
        print(cell.text())
        print(row)
        print(column)

        if cell.text() == '':
            return

        lr = 'None'
        region = cell.text()

        wb = load_workbook('regions.xlsx')
        sheet = wb.active

        column_regions = sheet['B']
        for index, cell_ex in enumerate(column_regions, start=1):
            if cell_ex.value == region:
                lr = sheet['A' + str(index)].value
                break
        wb.close()
        print(lr)
        self.ui.tableWidget.setItem(row, column, QTableWidgetItem(str(lr)))

    def get_gl_regions(self):
        wb = load_workbook('google_regions.xlsx')
        sheet = wb.active
        regions_name = sheet['A']
        uule_code = sheet['B']

        # проверить
        regions = {a.value: b.value for a, b in zip(regions_name, uule_code)}

        wb.close()

        return regions

    # Функция возвращает список с городами для пользовательского поиска
    def get_yd_regions(self):
        wb = load_workbook('regions.xlsx')
        sheet = wb.active
        column_regions = sheet['B']
        # переделать в словарь и затем далее по коду
        regions = tuple(cell.value for cell in column_regions)
        print(regions)
        wb.close()
        return regions

    # Функция возвращает словарь регионов из виджета
    def get_regions_from_table(self):
        regions = {}
        for row in range(0, 5):
            try:
                # Пропускаем, если пользователь в таблице оставил пустые строки или города с None
                if self.ui.tableWidget.cellWidget(row, 0).text() == '' or self.ui.tableWidget.item(row, 1).text() == 'None':
                    continue
                city = self.ui.tableWidget.cellWidget(row, 0).text()
                lr = self.ui.tableWidget.item(row, 1).text()
                regions[city] = lr
            except:
                pass

        if len(regions.items()) == 0:
            regions['current'] = ''

        return regions

    def get_regions_from_gl_table(self):
        regions = {}
        for row in range(0, 5):
            try:
                if self.ui.tableWidget_Google.cellWidget(row, 0).text() == '':
                    continue
                city = self.ui.tableWidget_Google.cellWidget(row, 0).text()
                uule = self.gl_regions[city]
                regions[city] = uule
            except:
                pass
        if len(regions.items()) == 0:
            regions['current'] = ''

        return regions

    # Функция для создания файла по экселю и записи в него статистики
    def edit_file_stat(self, statistics, main_folder_path):
        # Разбиваем статистику на два массива яндекса и гугла
        yd_statistics = []
        gl_statistics = []
        for stat in statistics:
            if stat[7] == 'yandex':
                yd_statistics.append(stat)
            elif stat[7] == 'google':
                gl_statistics.append(stat)

        # Открываем шаблон файл экселя для записи статистики
        wb = load_workbook('template.xlsx')
        sheet = wb.active

        start = 'A3'
        end = f'G{len(yd_statistics)+3}'

        for cellObj, stat in zip(sheet[start:end], yd_statistics):
            for index, (cell, s) in enumerate(zip(cellObj, stat)):
                # 0 - Регион
                # 2 - Сайт
                if index == 0 or index == 2:
                    cell.value = s

                # 1 - Запрос. Если результатов нет, то красим ячейку в красный
                elif index == 1:
                    if stat[6] == 'Результатов нет':
                        cell.fill = PatternFill(start_color='da9694', fill_type='solid')
                    cell.value = s

                # 3, 4, 5 - Значения позиций на странице. Если объявления нет, то ставит прочерк
                elif index == 3 or index == 4 or index == 5:
                    if s[1] == 0:
                        cell.value = '-'
                    else:
                        cell.value = s[1]

                # 6 - Гиперссылка на скриншот
                elif index == 6:
                    cell.value = s
                    cell.hyperlink = s
                    cell.style = 'Hyperlink'

        start = 'H3'
        end = f'N{len(gl_statistics)+3}'

        for cellObj, stat in zip(sheet[start:end], gl_statistics):
            for index, (cell, s) in enumerate(zip(cellObj, stat)):
                # 0 - Регион
                # 2 - Сайт
                if index == 0 or index == 2:
                    cell.value = s

                # 1 - Запрос. Если результатов нет, то красим ячейку в красный
                elif index == 1:
                    if stat[6] == 'Результатов нет':
                        cell.fill = PatternFill(start_color='da9694', fill_type='solid')
                    cell.value = s

                # 3, 4, 5 - Значения позиций на странице. Если объявления нет, то ставит прочерк
                elif index == 3 or index == 4 or index == 5:
                    if s[1] == 0:
                        cell.value = '-'
                    else:
                        cell.value = s[1]

                # 6 - Гиперссылка на скриншот
                elif index == 6:
                    cell.value = s
                    cell.hyperlink = s
                    cell.style = 'Hyperlink'

        wb.save('{0}\\statistics.xlsx'.format(main_folder_path))
        wb.close()

    # Открывает папку, куда указал пользователь
    def open_folder(self, main_folder_path):
        main_folder_path = main_folder_path.replace('/', '\\')
        os.system('explorer "{0}"'.format(main_folder_path))

    # Открывает excel файл со статистикой
    def open_excel_file(self, main_folder_path):
        main_folder_path = main_folder_path.replace('/', '\\')
        os.system('explorer "{0}\\statistics.xlsx"'.format(main_folder_path))

    def start_search(self):
        self.thread_instance.start()

    def end_search(self):
        self.thread_instance.stop()
        self.ui.label_Info.setText('Поиск остановлен')

    # Функция срабатывает при нажатии на кнопку Старт
    def start_searching(self):
        self.ui.pushButton_Start.setDisabled(True)
        self.ui.label_Info.setText('Программа выполняется. Подождите...')

        # Адреса сайтов, который ввел пользователь
        sites_addresses = self.get_sites_addresses(self.ui.textEdit_SitesAddresses.toPlainText())
        print(sites_addresses)

        # Массив с запросами для поиска
        user_requests = self.get_requests(self.ui.textEdit_Requests.toPlainText())
        print(user_requests)

        # Словарь с регионами и кодами lr для Яндекса
        yd_regions = self.get_regions_from_table()
        print(yd_regions)

        # Словарь с регионами и кодами для Google
        gl_regions = self.get_regions_from_gl_table()
        print(gl_regions)

        # Проверка ввел ли пользователь все данные
        if len(user_requests) == 0:
            self.ui.label_Info.setText('Ошибка! Вы не указали запросы!')
            self.thread_instance.stop()
            return ''
        elif len(sites_addresses) == 0 and not self.ui.radioButton_SpecialAndGarant.isChecked():
            self.ui.label_Info.setText('Ошибка! Вы не указали сайты!')
            self.thread_instance.stop()
            return ''
        elif self.save_path is None or self.save_path == '':
            self.ui.label_Info.setText('Ошибка! Вы не указали путь для сохранения!')
            self.thread_instance.stop()
            return ''

        print(len(user_requests))
        print(len(yd_regions))
        print(self.searchers)
        print(self.save_path)

        # Путь для создания папки со скриншотами
        folder_name=''
        if len(sites_addresses) == 0:
            folder_name = 'adhunter'
        else:
            folder_name = sites_addresses[0]

        main_folder_path = '{0}\\AH_{1}_{2}'.format(self.save_path, folder_name, datetime.now().strftime('%H_%M_%S'))
        self.ui.label_SavePath.setText('Путь: {0}'.format(main_folder_path))
        if not os.path.exists(main_folder_path):
            os.mkdir(main_folder_path)

        # Массив для сбора статистики
        statistics = []

        # Настройки браузера
        options = webdriver.ChromeOptions()

        # Оконный режим
        if self.ui.radioButton_Windowscreen.isChecked():
            options.add_argument('start-maximized')
            print('браузер открыт')
        # Окно скрыто,скрин полный
        elif self.ui.radioButton_Fullscreen.isChecked():
            options.add_argument('window-size=1920x3800')
            options.add_argument('headless')
            print('браузер скрыт')
        elif self.ui.radioButton_OnlyAd.isChecked():
            options.add_argument('start-maximized')
            print('браузер открыт')
        elif self.ui.radioButton_SpecialAndGarant.isChecked():
            options.add_argument('start-maximized')
            print('браузер открыт')

        if self.ui.checkBox_RotateScreen.isChecked():
            proc = subprocess.Popen(['powershell.exe', "./co.cmd"])
            proc.wait()
        print('???1')
        # Открываем браузер с заданными настройками
        driver = webdriver.Chrome(options=options)  # options=options
        print('???2')
        # Перебор всех поисковых систем
        for url in self.searchers:
            regions = {}
            search = ''
            padding = 0
            if url == 'https://yandex.ru/search/?text={0}&lr={1}':
                search = 'yandex'
                regions = yd_regions
                padding = 100 #При скролле к объявлению
                x_padding = 116 #При рисовании прямоугольника
                y_padding = 95 #При рисовании прямоугольника
                search_size = 641
            elif url == 'https://www.google.com/search?q={0}&uule={1}':
                search = 'google'
                regions = gl_regions
                padding = 60 #При скролле к объявлению
                x_padding = 180 #150 #При рисовании прямоугольника
                y_padding = 55 #При рисовании прямоугольника
                search_size = 645

            # Перебор регионов, по которым ведётся поиск
            for region, code in regions.items():
                print(f'region={region}')
                print(code)

                # Перебор всех поисковых запросов
                for request in user_requests:
                    current_url = url.format(request, code)
                    driver.get(current_url)
                    print(f'current_url={current_url}')

                    # Находим все элементы выдачи на странице
                    if search == 'yandex':
                        web_results = driver.find_elements_by_xpath('//li[@class="serp-item"]')
                        # Убираем рекламу Яндекс.Маркета
                        for result in web_results:
                            if 'Яндекс.Маркет' in result.text and 'Реклама' in result.text:
                                web_results.remove(result)
                                break
                    elif search == 'google':
                        # Здесь сделано таким образом, потмому что в поиске гугла у рекламных элементов разная вложенность.
                        # Пока что встречал два разных варианта.
                        top_results = driver.find_elements_by_xpath('//*[@id="tads"]/div/ol/li')
                        if not len(top_results) == 0:
                            garant_results = driver.find_elements_by_xpath('//*[@id="tadsb"]/div/ol/li')
                        elif len(top_results) == 0:
                            top_results = driver.find_elements_by_xpath('//*[@id="tads"]/div')
                            garant_results = driver.find_elements_by_xpath('//*[@id="tadsb"]/div')
                        seo_results = driver.find_elements_by_xpath('//div[@class="rc"]')
                        print('блоки')
                        print(top_results)
                        print(garant_results)
                        web_results = top_results + seo_results + garant_results

                    # Новый режим, только спец и гарант без конкретного сайта
                    if self.ui.radioButton_SpecialAndGarant.isChecked():
                        results = []
                        self.ui.label_Info.setText(
                            'Выполняю запрос: {0} - {1}'.format(region, request))
                        print('Выполняю запрос: {0} - {1}'.format(region, request))

                        # Собираем всю инфу со страницы
                        for index, result in enumerate(web_results, start=1):
                            if 'Яндекс.Маркет' in result.text and 'Реклама' in result.text:
                                continue
                            results.append((index, result.text, result.location, result.size))

                        block_of_ads = self.get_block_of_ads(results)
                        print(block_of_ads[0])
                        print(block_of_ads[2])
                        # Проверяем наличие spec
                        if len(block_of_ads[0]) != 0:

                            if not os.path.exists(f'{main_folder_path}\\{search}\\{region}'):
                                os.makedirs(f'{main_folder_path}\\{search}\\{region}')

                            screen_name = f'{main_folder_path}\\{search}\\{region}\\{search}_{region}_{request}_special_screen.png'
                            print(screen_name)

                            # driver.save_screenshot(screen_name)

                            sleep(1)
                            # скрин
                            img = ImageGrab.grab().save(screen_name, 'PNG')
                            img = Image.open(screen_name)

                            new_img = img.crop((0, 120, img.width, img.height))
                            print('tut')
                            # Отрезает всё что ниже искомого объявления, оставляет только поисковую
                            # строку и само объявление
                            # if search == 'google':
                            #     print('или тут')
                            #     new_img = new_img.crop(
                            #         (0, 0, x_padding + search_size,
                            #          block_of_ads[0][-1][3]['height'] + block_of_ads[0][-1][2]['y']))
                            #     print('или тут')
                            # else:
                            new_img = new_img.crop(
                                (0, 0, x_padding + search_size + 50, block_of_ads[0][-1][3]['height'] + block_of_ads[0][-1][2]['y']))
                            new_img.save(screen_name, 'PNG')
                            if search == 'google':
                                print('сделал скрин для гугла')

                        # Проверяем на garant
                        if len(block_of_ads[2]) != 0:

                            if not os.path.exists(f'{main_folder_path}\\{search}\\{region}'):
                                os.makedirs(f'{main_folder_path}\\{search}\\{region}')

                            screen_name = f'{main_folder_path}\\{search}\\{region}\\{search}_{region}_{request}_garant_screen.png'
                            print(screen_name)

                            # driver.save_screenshot(screen_name)
                            driver.execute_script(
                                "window.scrollTo(0, {0})".format(block_of_ads[2][0][2]['y'] - padding))
                            sleep(1)
                            # скрин

                            img = ImageGrab.grab().save(screen_name, 'PNG')
                            img = Image.open(screen_name)
                            #
                            new_img = img.crop((0, 120, img.width, img.height-95))
                            print('tut')
                            # Отрезает всё что ниже искомого объявления, оставляет только поисковую
                            # строку и само объявление
                            # if search == 'google':
                            #     print('или тут')
                            #     # new_img = new_img.crop(
                            #     #     (0, 0, x_padding + search_size,
                            #     #      block_of_ads[2][-1][3]['height'] + block_of_ads[2][-1][2]['y']))
                            #     new_img = new_img.crop(
                            #         (0, 0, x_padding + search_size, new_img.size[1]))
                            #     print('или тут')
                            # else:
                            new_img = new_img.crop(
                                    (0, 0, x_padding + search_size + 50, new_img.size[1]))
                            new_img.save(screen_name, 'PNG')

                            # для яндекса готово. осталось затестить для гугла
                    else:
                    # Находим все адреса сайтов в выдаче
                    # sites_on_page = driver.find_elements_by_xpath('//li[@class="serp-item"]/div/div/div/a/b')
                        for site_address in sites_addresses:
                            self.ui.label_Info.setText('Выполняю запрос: {0} - {1} - {2}'.format(region, request, site_address))
                            results = []

                            # Лист для позиций искомого сайта в выдаче
                            # Первое число позиция по всем запросам, второе относительно блока в котором находится
                            positions = [(0, 0), (0, 0), (0, 0)]
                            screen_name = 'Результатов нет'

                            # TODO в каждом результате много данных и в них программа ищет наличие сайта
                            # TODO для оптимизации следует собирать со страницы только адреса, сравнивать их с нашим сайтом
                            # Перебор результатов выдачи поиска
                            # Режим оконного скрина
                            if self.ui.radioButton_Windowscreen.isChecked():
                                for index, r in enumerate(web_results, start=1):
                                    if 'Яндекс.Маркет' in r.text and 'Реклама' in r.text:
                                        continue
                                    results.append((index, r.text, r.location, r.size))
                                positions, block_of_ads = self.get_positions(results, site_address)

                                # Только спец размещение
                                for i in range(0, 4):
                                    if site_address in web_results[i].text and ('реклама' in web_results[i].text or 'Реклама' in web_results[i].text):
                                        print(web_results[i].text)

                                        # Определяем прокручивать к скрину или нет
                                        if not self.ui.checkBox_WithoutScrollDown.isChecked():
                                            if i == 0 and search == 'google':
                                                driver.execute_script("window.scrollTo(0, 0)")
                                            else:
                                                driver.execute_script(
                                                    "window.scrollTo(0, {0})".format(web_results[i].location['y'] - padding))

                                        if not os.path.exists(f'{main_folder_path}\\{search}\\{site_address}\\{region}'):
                                            os.makedirs(f'{main_folder_path}\\{search}\\{site_address}\\{region}')

                                        screen_name = f'{main_folder_path}\\{search}\\{site_address}\\{region}\\{search}_{region}_{request}_{site_address}_window_screen.png'
                                        print(screen_name)
                                        sleep(1)
                                        #скрин
                                        img = ImageGrab.grab().save(screen_name, 'PNG')
                                        img = Image.open(screen_name)

                                        # Отрезает верхнюю часть, где находится адресная строка, оставляет всё что ниже
                                        new_img = img.crop((0, 120, img.width, img.height))

                                        if not self.ui.checkBox_WithoutFrame.isChecked():
                                            # Рисуем рамку
                                            # Скроллит к объявлению
                                            if not self.ui.checkBox_WithoutScrollDown.isChecked():
                                                self.ui.label_Info.setText(
                                                    'Выполняю запрос: {0} - {1} - {2}. Рисую на скрине'.format(region, request,
                                                                                                               site_address))
                                                draw = ImageDraw.Draw(new_img)

                                                if i == 0 and search == 'google':
                                                        draw.rectangle((web_results[i].location['x'] - 2, web_results[i].location['y'] - 2,
                                                                        web_results[i].size['width'] + web_results[i].location['x'],
                                                                        web_results[i].size['height'] + web_results[i].location['y']),
                                                                       outline=(255, 0, 0, 255),
                                                                       width=2)
                                                else:
                                                    draw.rectangle((x_padding, 0 + y_padding,
                                                                    web_results[i].size['width'] + x_padding,
                                                                    web_results[i].size['height'] + y_padding),
                                                                   outline=(255, 0, 0, 255),
                                                                   width=2)

                                            # Без скролла к объявлению
                                            elif self.ui.checkBox_WithoutScrollDown.isChecked():
                                                self.ui.label_Info.setText(
                                                    'Выполняю запрос: {0} - {1} - {2}. Рисую на скрине'.format(region,
                                                                                                               request,
                                                                                                               site_address))
                                                draw = ImageDraw.Draw(new_img)

                                                draw.rectangle(
                                                    (web_results[i].location['x'] - 2, web_results[i].location['y'] - 2,
                                                     web_results[i].size['width'] + web_results[i].location['x'],
                                                     web_results[i].size['height'] + web_results[i].location['y']),
                                                    outline=(255, 0, 0, 255),
                                                    width=2)

                                        # Рисуем нумерацию
                                        font = ImageFont.truetype('Aegean.ttf', 25)
                                        if self.ui.checkBox_WithoutScrollDown.isChecked() and self.ui.checkBox_Numeration.isChecked():
                                            for result in results:
                                                draw.text((result[2]['x'] - 100, result[2]['y']), str(result[0]),
                                                          fill=(0, 0, 0), font=font)

                                            for index, spcl in enumerate(block_of_ads[0], start=1):
                                                draw.text((spcl[2]['x'] - 50, spcl[2]['y']), str(index),
                                                          fill=(255, 0, 0), font=font)

                                            for index, s in enumerate(block_of_ads[1], start=1):
                                                draw.text((s[2]['x'] - 50, s[2]['y']), str(index), fill=(0, 128, 0),
                                                          font=font)

                                            for index, g in enumerate(block_of_ads[2], start=1):
                                                draw.text((g[2]['x'] - 50, g[2]['y']), str(index), fill=(0, 0, 255),
                                                          font=font)

                                        # Тут идет смещение нумерации. Всю просто нумерацию выводить неправильно
                                        #
                                        elif not self.ui.checkBox_WithoutScrollDown.isChecked() and self.ui.checkBox_Numeration.isChecked():
                                            results = results[i:]

                                            for result in results:
                                                if result[0] != 1 and search != 'google':
                                                    y = result[2]['y'] - results[0][2]['y'] + y_padding
                                                else:
                                                    y = result[2]['y']

                                                draw.text((result[2]['x'] - 100, y), str(result[0]),
                                                          fill=(0, 0, 0), font=font)

                                        new_img.save(screen_name, 'PNG')
                                        break

                            # Режим полного скрина
                            elif self.ui.radioButton_Fullscreen.isChecked():
                                print('фулл скрин')
                                for result in web_results:
                                    if site_address in result.text:
                                        print(result.text)

                                        # Собираем всю инфу со страницы
                                        for index, result in enumerate(web_results, start=1):
                                            if 'Яндекс.Маркет' in result.text and 'Реклама' in result.text:
                                                continue
                                            results.append((index, result.text, result.location, result.size))

                                        if not os.path.exists(f'{main_folder_path}\\{search}\\{site_address}\\{region}'):
                                            os.makedirs(f'{main_folder_path}\\{search}\\{site_address}\\{region}')

                                        # TODO тут с именем скриншота. Как его потом вытаскивать из функций
                                        screen_name = f'{main_folder_path}\\{search}\\{site_address}\\{region}\\{search}_{region}_{request}_{site_address}_full_screen.png'
                                        print(screen_name)

                                        driver.save_screenshot(screen_name)

                                        positions, block_of_ads = self.get_positions(results, site_address)

                                        # Рисуем на скрине
                                        self.ui.label_Info.setText('Выполняю запрос: {0} - {1} - {2}. Рисую на скрине'.format(region, request, site_address))
                                        self.edit_screen(screen_name, results, positions, block_of_ads)
                                        break

                            # Режим только одной рекламы
                            elif self.ui.radioButton_OnlyAd.isChecked():
                                for index, r in enumerate(web_results, start=1):
                                    if 'Яндекс.Маркет' in r.text and 'Реклама' in r.text:
                                        continue
                                    results.append((index, r.text, r.location, r.size))
                                positions, block_of_ads = self.get_positions(results, site_address)

                                for i in range(0, len(web_results)):
                                    if site_address in web_results[i].text and (
                                            'реклама' in web_results[i].text or 'Реклама' in web_results[i].text):
                                        print(web_results[i].text)
                                        if i == 0 and search == 'google':
                                            driver.execute_script("window.scrollTo(0, 0)")
                                        else:
                                            driver.execute_script("window.scrollTo(0, {0})".format(web_results[i].location['y'] - padding))

                                        if not os.path.exists(f'{main_folder_path}\\{search}\\{site_address}\\{region}'):
                                            os.makedirs(f'{main_folder_path}\\{search}\\{site_address}\\{region}')

                                        screen_name = f'{main_folder_path}\\{search}\\{site_address}\\{region}\\{search}_{region}_{request}_{site_address}_only_ad_screen.png'
                                        print(screen_name)
                                        sleep(1)
                                        # скрин
                                        img = ImageGrab.grab().save(screen_name, 'PNG')
                                        img = Image.open(screen_name)

                                        # Отрезает верхнюю часть, где находится адресная строка, оставляет всё что ниже
                                        new_img = img.crop((0, 120, img.width, img.height))

                                        # Отрезает всё что ниже искомого объявления, оставляет только поисковую
                                        # строку и само объявление
                                        if i == 0 and search == 'google':
                                            new_img = new_img.crop(
                                                (0, 0, x_padding + search_size, web_results[i].size['height'] + web_results[i].location['y']))
                                        else:
                                            new_img = new_img.crop((0, 0, x_padding + search_size, web_results[i].size['height'] + padding))
                                        new_img.save(screen_name, 'PNG')
                                        break

                            # Разбиваем массив с позициями на несколько массивов, так проще потом обрабатывать
                            spec = positions[0]
                            seo = positions[1]
                            garant = positions[2]

                            # Добавляем данные в один большой, чтобы потом записать всё в эксель файл
                            statistics.append((region, request, site_address, spec, seo, garant, screen_name, search))

        # Закрываем браузер
        driver.close()
        print('закрыл драйвер')

        if self.ui.checkBox_RotateScreen.isChecked():
            proc = subprocess.Popen(['powershell.exe', "./co.cmd"])
            proc.wait()

        # Вот тут желательно перебрать скрины и обработать их
        self.ui.label_Info.setText('Обрабатываю скриншоты')
        # "тут"

        # Записываем статистику в файл
        if not self.ui.radioButton_SpecialAndGarant.isChecked():
            self.ui.label_Info.setText('Собираю статистику в файл')
            self.edit_file_stat(statistics, main_folder_path)
            print('записал стату в файл')

        # Закрывает процесс chromedriver
        os.system("TASKKILL /F /IM chromedriver.exe")
        print('закрыл chromedriver')

        self.ui.label_Info.setText('Готово! ( ͡° ͜ʖ ͡°)')
        # Проверка указал ли пользователь открывать папку
        if self.ui.checkBox_OpenFolder.isChecked():
            self.open_folder(main_folder_path)
            print('открыл папку')

        # Проверка указал ли пользователь открывать excel файл
        if self.ui.checkBox_OpenExcelFile.isChecked():
            self.open_excel_file(main_folder_path)
            print('открыл файл')

        self.thread_instance.stop()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    myApp = MyWin()
    myApp.show()
    sys.exit(app.exec_())
