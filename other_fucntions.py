import os
import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# Функция разбивает текст пользователя в textEdit "Запросы" в массив.
def get_requests(text):
    temp = text.split('\n')
    user_requests = tuple(request for request in temp if request != '')
    return user_requests


# Функция разбивает текст пользователя в textEdit "Сайты" в массив.
def get_sites_addresses(text):
    clear_sites_addresses = []
    pattern = r'\w+-*\w+\.\w+-*\w+\.*\w*'
    sites_addresses = text.split('\n')
    print(sites_addresses)

    for site in sites_addresses:
        match = re.search(pattern, site)
        if match:
            match = match[0]
            if match[0:4] == 'www.':
                match = match[4:]
            clear_sites_addresses.append(match)
    print(clear_sites_addresses)
    clear_sites_addresses = tuple(site for site in clear_sites_addresses)
    return clear_sites_addresses


# Функция разбивает все результаты поиска на три массива спец/сео/гарант.
# Возвращает массив массивов.
def get_block_of_ads(results):
    special = []
    seo = []
    guaranteed = []

    temp = False

    # Распределяем результаты по массивам
    for result in results:
        if 'Яндекс.Маркет' in result[1] and 'Реклама' in result[1]:
            continue
        if 'реклама' in result[1] or 'Реклама' in result[1]:
            if temp:
                guaranteed.append(result)
            else:
                special.append(result)
        else:
            temp = True
            seo.append(result)

    block_of_ads = [special, seo, guaranteed]
    return block_of_ads


# Функция вычисляет позицию сайта относительно блоков спец/сео/гарант
def get_site_position(block, site_address):
    count = 0
    index = 0
    for index, b in enumerate(block, start=1):
        if site_address in b[1]:
            count = b[0]
            break
    if count == 0:
        index = 0

    return count, index


# Функция для определения позиций в спец/seo/гарант.
def get_positions(self, results, site_address=None):
    special = []
    seo = []
    guaranteed = []

    temp = False

    # Распределяем результаты по массивам
    for result in results:
        if 'реклама' in result[1] or 'Реклама' in result[1]:
            if temp:
                guaranteed.append(result)
            else:
                special.append(result)
        else:
            temp = True
            seo.append(result)

    special_position = self.get_site_position(special, site_address)
    seo_position = self.get_site_position(seo, site_address)
    guaranteed_position = self.get_site_position(guaranteed, site_address)

    positions = [special_position, seo_position, guaranteed_position]
    block_of_ads = [special, seo, guaranteed]
    return positions, block_of_ads


# Считывает из файла регионы гугла.
# Возвращает словарь {название региона: код_для_урла}
def get_gl_regions():
    wb = load_workbook('excel/google_regions.xlsx')
    sheet = wb.active
    regions_name = sheet['A']
    uule_code = sheet['B']

    regions = {a.value: b.value for a, b in zip(regions_name, uule_code)}

    wb.close()

    return regions


# Функция возвращает список с городами для пользовательского поиска.
# Считывает регионы для яндекса из файла.
# Возвращает кортеж.
def get_yd_regions():
    wb = load_workbook('excel/regions.xlsx')
    sheet = wb.active
    column_regions = sheet['B']
    regions = tuple(cell.value for cell in column_regions)
    wb.close()
    return regions


# Функция для создания файла экселя и записи в него статистики.
def edit_file_stat(statistics, main_folder_path):
    # Разбиваем статистику на два массива яндекса и гугла.
    yd_statistics = []
    gl_statistics = []
    for stat in statistics:
        if stat[7] == 'yandex':
            yd_statistics.append(stat)
        elif stat[7] == 'google':
            gl_statistics.append(stat)

    # Открываем шаблон файл экселя для записи статистики.
    wb = load_workbook('excel/template.xlsx')
    sheet = wb.active

    start = 'A3'
    end = f'G{len(yd_statistics) + 3}'

    # Цикл для статистики по Яндексу.
    for cellObj, stat in zip(sheet[start:end], yd_statistics):
        for index, (cell, s) in enumerate(zip(cellObj, stat)):
            # 0 - Регион
            # 2 - Сайт
            if index == 0 or index == 2:
                cell.value = s

            # 1 - Запрос. Если результатов нет, то красим ячейку в красный.
            elif index == 1:
                if stat[6] == 'Результатов нет':
                    cell.fill = PatternFill(start_color='da9694', fill_type='solid')
                cell.value = s

            # 3, 4, 5 - Значения позиций на странице. Если объявления нет, то ставит прочерк.
            elif index == 3 or index == 4 or index == 5:
                if s[1] == 0:
                    cell.value = '-'
                else:
                    cell.value = s[1]

            # 6 - Гиперссылка на скриншот
            elif index == 6:
                cell.value = s
                cell.hyperlink = s
                cell.style = 'Hyperlink'

    start = 'H3'
    end = f'N{len(gl_statistics) + 3}'

    # Цикл для статистики по Гуглу.
    for cellObj, stat in zip(sheet[start:end], gl_statistics):
        for index, (cell, s) in enumerate(zip(cellObj, stat)):
            # 0 - Регион
            # 2 - Сайт
            if index == 0 or index == 2:
                cell.value = s

            # 1 - Запрос. Если результатов нет, то красим ячейку в красный.
            elif index == 1:
                if stat[6] == 'Результатов нет':
                    cell.fill = PatternFill(start_color='da9694', fill_type='solid')
                cell.value = s

            # 3, 4, 5 - Значения позиций на странице. Если объявления нет, то ставит прочерк.
            elif index == 3 or index == 4 or index == 5:
                if s[1] == 0:
                    cell.value = '-'
                else:
                    cell.value = s[1]

            # 6 - Гиперссылка на скриншот
            elif index == 6:
                cell.value = s
                cell.hyperlink = s
                cell.style = 'Hyperlink'

    wb.save('{0}\\statistics.xlsx'.format(main_folder_path))
    wb.close()


# Открывает папку, куда указал пользователь.
def open_folder(main_folder_path):
    main_folder_path = main_folder_path.replace('/', '\\')
    os.system('explorer "{0}"'.format(main_folder_path))


# Открывает excel файл со статистикой.
def open_excel_file(main_folder_path):
    main_folder_path = main_folder_path.replace('/', '\\')
    os.system('explorer "{0}\\statistics.xlsx"'.format(main_folder_path))
